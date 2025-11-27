import asyncio
import sys
from pathlib import Path
from typing import List

from bson import ObjectId
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from app.database import init_db  # type: ignore
from app.models.user import User  # type: ignore

EXCEL_PATH = Path(r"E:\interships 2026 startover\Shift\ShiftSync\new_employees.xlsx")


def load_employee_ids() -> List[ObjectId]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: i for i, header in enumerate(headers)}

    ids: List[ObjectId] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["role"]] != "employee":
            continue
        raw = row[idx["_id"]]
        oid = str(raw).replace("ObjectId('", "").replace("')", "")
        ids.append(ObjectId(oid))
    return ids


async def disable_employees() -> None:
    employee_ids = load_employee_ids()
    if not employee_ids:
        print("No employee IDs found in Excel; nothing to disable.")
        return

    await init_db()

    users = await User.find({"_id": {"$in": employee_ids}}).to_list()
    updated = 0
    missing = 0

    for user in users:
        if user.status != "disabled":
            user.status = "disabled"
            await user.save()
            updated += 1
    missing = len(employee_ids) - len(users)

    print(f"Found {len(users)} matching employees; disabled {updated}. Missing: {missing}.")


if __name__ == "__main__":
    asyncio.run(disable_employees())
