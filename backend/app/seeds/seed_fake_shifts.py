import asyncio
import calendar
import sys
from datetime import date
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple

from bson import ObjectId
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from app.database import init_db  # type: ignore
from app.models.shift import Shift  # type: ignore

EMPLOYEE_SOURCE = Path(r"E:\interships 2026 startover\Shift\ShiftSync\shiftsync_employees_final.xlsx")

SHIFT_WINDOWS: Tuple[Tuple[str, str], Tuple[str, str]] = (("09:00", "17:00"), ("13:00", "21:00"))
COMPLETED_MONTHS = [(2025, 9), (2025, 11)]
ASSIGNED_MONTHS = [(2025, 12)]
HOLIDAYS = {date(2025, 9, 1), date(2025, 12, 25), date(2025, 12, 26)}


def load_active_employees() -> List[ObjectId]:
    wb = load_workbook(EMPLOYEE_SOURCE, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: i for i, header in enumerate(headers)}

    employees: List[ObjectId] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["role"]] != "employee":
            continue
        if row[idx["status"]] != "active":
            continue
        raw = row[idx["_id"]]
        oid = str(raw).replace("ObjectId('", "").replace("')", "")
        employees.append(ObjectId(oid))
    return employees


def month_days(year: int, month: int) -> Iterable[date]:
    _, last_day = calendar.monthrange(year, month)
    for day in range(1, last_day + 1):
        yield date(year, month, day)


def build_day_assignments(employees: Sequence[ObjectId], day_index: int) -> List[Tuple[ObjectId, str, str]]:
    total = len(employees)
    start = day_index % total
    morning_indices = [(start + i) % total for i in range(10)]
    afternoon_indices = [(start + 10 + i) % total for i in range(10)]

    assignments: List[Tuple[ObjectId, str, str]] = []
    for idx in morning_indices:
        assignments.append((employees[idx], SHIFT_WINDOWS[0][0], SHIFT_WINDOWS[0][1]))
    for idx in afternoon_indices:
        assignments.append((employees[idx], SHIFT_WINDOWS[1][0], SHIFT_WINDOWS[1][1]))
    return assignments


async def seed_shifts() -> None:
    employees = load_active_employees()
    if len(employees) < 20:
        raise RuntimeError(f"Need at least 20 active employees, found {len(employees)}")

    print(f"Loaded {len(employees)} active employees from {EMPLOYEE_SOURCE.name}")
    await init_db()

    to_insert: List[Shift] = []
    day_counter = 0
    planned: List[Tuple[str, date]] = []

    for year, month in COMPLETED_MONTHS:
        for shift_day in month_days(year, month):
            if shift_day in HOLIDAYS:
                continue
            for emp_id, start, end in build_day_assignments(employees, day_counter):
                to_insert.append(
                    Shift(
                        employee_id=emp_id,
                        shift_date=shift_day,
                        start_time=start,
                        end_time=end,
                        status="completed",
                    )
                )
            planned.append(("completed", shift_day))
            day_counter += 1

    for year, month in ASSIGNED_MONTHS:
        for shift_day in month_days(year, month):
            if shift_day in HOLIDAYS:
                continue
            for emp_id, start, end in build_day_assignments(employees, day_counter):
                to_insert.append(
                    Shift(
                        employee_id=emp_id,
                        shift_date=shift_day,
                        start_time=start,
                        end_time=end,
                        status="assigned",
                    )
                )
            planned.append(("assigned", shift_day))
            day_counter += 1

    if not to_insert:
        print("Nothing to insert")
        return

    completed_days = len({d for status, d in planned if status == "completed"})
    assigned_days = len({d for status, d in planned if status == "assigned"})
    print(f"Planned {len(to_insert)} shifts: {completed_days} days completed, {assigned_days} days assigned")

    result = await Shift.insert_many(to_insert)
    inserted_count = len(result.inserted_ids) if hasattr(result, "inserted_ids") else len(to_insert)
    print(f"Inserted {inserted_count} shifts into shiftsync.shifts")


if __name__ == "__main__":
    asyncio.run(seed_shifts())
