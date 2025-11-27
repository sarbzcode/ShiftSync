from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from bson import ObjectId
from datetime import datetime, timezone
from io import BytesIO
from typing import List, Literal, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.models.user import User
from app.models.shift import Shift
from app.models.attendance import Attendance
from app.models.deleted_employee import DeletedEmployee
from app.schemas.user import (
    UserCreate,
    UserResponse,
    PasswordReset,
    UserUpdate,
    EmployeeSummary,
    EmployeeSearchResult,
)
from app.utils.security import hash_password
from app.utils.deps import require_admin, get_current_user
from app.services.system_settings import get_system_timezone

router = APIRouter()
EXPORT_HEADERS = ["Sr. No.", "Full Name", "Username", "Email", "Pay Rate"]


async def _gather_employee_rows() -> List[Tuple[int, str, str, str, float]]:
    """Fetch employee data for export."""
    employees = await User.find(User.role == "employee").sort("+name").to_list()
    rows: List[Tuple[int, str, str, str, float]] = []
    for index, employee in enumerate(employees, start=1):
        rows.append(
            (
                index,
                employee.name,
                employee.username,
                employee.email,
                float(employee.pay_rate),
            )
    )
    return rows


def _build_shift_datetime(shift_date, time_str: str, tz) -> datetime:
    hours, minutes = map(int, time_str.split(":"))
    return datetime(shift_date.year, shift_date.month, shift_date.day, hours, minutes, tzinfo=tz)


async def _latest_shift_clock_out(employee_id) -> Optional[datetime]:
    """Return the latest completed shift end time (timezone-aware UTC)."""
    shift = await Shift.find(
        Shift.employee_id == employee_id,
        Shift.status == "completed",
    ).sort("-shift_date", "-end_time").to_list(1)

    if not shift:
        return None

    tz = await get_system_timezone()
    end_dt = _build_shift_datetime(shift[0].shift_date, shift[0].end_time, tz)
    return end_dt.astimezone(timezone.utc)


def _build_excel_workbook(rows: List[Tuple[int, str, str, str, float]]) -> BytesIO:
    """Create an Excel file in memory."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Employees"
    worksheet.append(EXPORT_HEADERS)

    for row in rows:
        sr_no, full_name, username, email, pay_rate = row
        worksheet.append([sr_no, full_name, username, email, pay_rate])

    for column_index, header in enumerate(EXPORT_HEADERS, start=1):
        column_letter = get_column_letter(column_index)
        max_length = len(header)
        for cell in worksheet[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _build_pdf_document(rows: List[Tuple[int, str, str, str, float]]) -> BytesIO:
    """Create a PDF table in memory."""
    stream = BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    table_data = [
        EXPORT_HEADERS,
        *[
            [
                str(sr_no),
                full_name,
                username,
                email,
                f"${pay_rate:,.2f}",
            ]
            for sr_no, full_name, username, email, pay_rate in rows
        ],
    ]

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4ed8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5f5")),
            ]
        )
    )

    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Employee Directory", styles["Heading2"]),
        Spacer(1, 12),
        table,
    ]
    doc.build(elements)
    stream.seek(0)
    return stream

def _serialize_user(user: User) -> UserResponse:
    return UserResponse(
        id=str(user.id),
        username=user.username,
        name=user.name,
        email=user.email,
        role=user.role,
        pay_rate=user.pay_rate,
        status=user.status,
        created_at=user.created_at
    )

@router.get("/me", response_model=UserResponse)
async def get_current_profile(current_user: User = Depends(get_current_user)):
    """Return the authenticated user's profile."""
    return _serialize_user(current_user)

@router.post("", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def create_user(user_data: UserCreate, admin: User = Depends(require_admin)):
    """Create a new employee user (Admin only)."""
    # Check if username or email already exists
    existing_user = await User.find_one(
        {"$or": [{"username": user_data.username}, {"email": user_data.email}]}
    )
    
    if existing_user:
        if existing_user.username == user_data.username:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Username already exists"
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Email already exists"
            )
    
    # Create new user
    new_user = User(
        username=user_data.username,
        password_hash=hash_password(user_data.password),
        role="employee",  # Always create as employee
        name=user_data.name,
        email=user_data.email,
        pay_rate=user_data.pay_rate,
        status="active"
    )
    
    await new_user.insert()
    
    return _serialize_user(new_user)

@router.get("", response_model=List[UserResponse])
async def list_users(admin: User = Depends(require_admin)):
    """List all users (Admin only)."""
    users = await User.find(User.role == "employee").to_list()
    
    return [_serialize_user(user) for user in users]

@router.get("/search", response_model=List[EmployeeSearchResult])
async def search_employees(
    q: str = Query(..., min_length=1, description="Name, username, or email search"),
    limit: int = Query(10, ge=1, le=50),
    admin: User = Depends(require_admin),
):
    """Search employees by name, username, or email (Admin only)."""
    query = q.strip()
    if not query:
        return []
    
    search_filter = {
        "role": "employee",
        "$or": [
            {"name": {"$regex": query, "$options": "i"}},
            {"username": {"$regex": query, "$options": "i"}},
            {"email": {"$regex": query, "$options": "i"}},
        ],
    }
    
    employees = await User.find(search_filter).limit(limit).to_list()
    
    return [
        EmployeeSearchResult(
            id=str(employee.id),
            username=employee.username,
            name=employee.name,
            email=employee.email,
            status=employee.status,
        )
        for employee in employees
    ]

@router.get("/management", response_model=List[EmployeeSummary])
async def list_employee_management(admin: User = Depends(require_admin)):
    """List employees with last clock-out information."""
    employees = await User.find(User.role == "employee").sort("-created_at").to_list()
    deleted_employees = await DeletedEmployee.find_all().sort("-deleted_at").to_list()
    results: List[EmployeeSummary] = []

    for employee in employees:
        last_clock = await Attendance.find(
            Attendance.user_id == employee.id,
            Attendance.clock_out != None  # type: ignore
        ).sort("-clock_out").to_list(1)
        last_attendance_out = last_clock[0].clock_out if last_clock else None

        last_shift_out = await _latest_shift_clock_out(employee.id)

        latest = None
        if last_attendance_out and last_shift_out:
            latest = max(last_attendance_out, last_shift_out)
        else:
            latest = last_attendance_out or last_shift_out

        results.append(
            EmployeeSummary(
                id=str(employee.id),
                username=employee.username,
                name=employee.name,
                email=employee.email,
                pay_rate=employee.pay_rate,
                status=employee.status,
                last_clock_out=latest,
                created_at=employee.created_at,
            )
        )

    for archived in deleted_employees:
        results.append(
            EmployeeSummary(
                id=str(archived.original_id),
                username=archived.username,
                name=archived.name,
                email=archived.email,
                pay_rate=archived.pay_rate,
                status="deleted",
                last_clock_out=None,
                created_at=archived.deleted_at,
            )
        )

    return results

@router.get("/export")
async def export_employees(
    format: Literal["excel", "pdf"] = Query(..., description="File format for employee export"),
    admin: User = Depends(require_admin),
):
    """Export employees as Excel or PDF."""
    rows = await _gather_employee_rows()
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    if format == "excel":
        stream = _build_excel_workbook(rows)
        filename = f"employees_{timestamp}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        stream = _build_pdf_document(rows)
        filename = f"employees_{timestamp}.pdf"
        media_type = "application/pdf"

    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(stream, media_type=media_type, headers=headers)

@router.put("/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: str,
    user_data: UserUpdate,
    admin: User = Depends(require_admin)
):
    """Update employee details (Admin only)."""
    try:
        user = await User.get(ObjectId(user_id))
    except Exception:
        user = None

    if not user or user.role != "employee":
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee not found"
        )

    update_values = user_data.model_dump(exclude_none=True)
    if not update_values:
        return _serialize_user(user)

    for field, value in update_values.items():
        setattr(user, field, value)

    await user.save()

    return _serialize_user(user)

@router.delete("/{user_id}", status_code=status.HTTP_200_OK)
async def delete_user(
    user_id: str,
    admin: User = Depends(require_admin)
):
    """Delete an employee and archive their information."""
    try:
        user = await User.get(ObjectId(user_id))
    except Exception:
        user = None

    if not user or user.role != "employee":
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee not found"
        )

    archived = DeletedEmployee(
        original_id=user.id,
        username=user.username,
        name=user.name,
        email=user.email,
        pay_rate=user.pay_rate,
        status=user.status,
    )
    await archived.insert()

    await user.delete()

    return {"message": "Employee deleted and archived"}

@router.put("/{user_id}/reset-password")
async def reset_password(
    user_id: str,
    password_data: PasswordReset,
    admin: User = Depends(require_admin)
):
    """Reset user password (Admin only)."""
    try:
        user = await User.get(ObjectId(user_id))
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Update password
    user.password_hash = hash_password(password_data.new_password)
    await user.save()
    
    return {"message": "Password reset successfully", "username": user.username}
